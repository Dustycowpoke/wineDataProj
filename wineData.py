# wineData.py - collects as much data as possible from LCBO's API

#***TODO***
# only returning 1 page

import requests
import json
import openpyxl
from wineData_pass import *

# LCBO API 
TOKEN_ACCESS_KEY = GITHUB_TOKEN_ACCESS_KEY
LCBO_URL = 'https://lcboapi.com/products'
results_per_page = 100

# Global wine score API 
GWS_URL = 'https://api.globalwinescore.com/globalwinescores/latest'
TOKEN = GITHUB_TOKEN 

# Global iterator for writing to spreadsheet
row_iter = 2

def lcbo_api():
	
	# Create workbook
	wb = openpyxl.Workbook()
	wb.title = 'wineData'
	sheet = wb.active
	
	# Create columns with headers
	sheet['A1'] = 'name'
	sheet['B1'] = 'price'
	
	'''******************************'''
	
	# Set payload to API
	payload = {'access_key':TOKEN_ACCESS_KEY,
				'per_page': results_per_page,
				'q':'bordeaux',
				'is_dead':'false'
				}
	
	# Pass payload to API then to text object
	page_1_obj = requests.get(LCBO_URL, payload)
	page_1_doc = json.loads(page_1_obj.text)
		
	# Find total pages in 'pager' section to iterate over
	page_1_pager = page_1_doc.get('pager')
	total_pages = page_1_pager.get('total_pages')
	wine_records_on_page = int(page_1_pager.get('current_page_record_count'))
	
	# Get results from page 1
	wine_results = page_1_doc.get('result')
	
	# Iterate over rows to write data to Excel
	global row_iter
		
	# For each record on page 1
	for j in range(0, (wine_records_on_page)):
		
		name = wine_results[j].get('name')
		price = int(wine_results[j].get('price_in_cents'))
					
		# Write data to next row
		sheet.cell(row = row_iter, column = 1).value = name
		sheet.cell(row = row_iter, column = 2).value = price
			
		# Increment row_iter
		row_iter += 1
	
	# Loop thru every subsequent page
	for page in range(2, (total_pages + 1)):
		
		payload['page'] = page
		page_obj = requests.get(LCBO_URL, payload)
		page_doc = json.loads(page_obj.text)
		
		# Get results on this page
		page_pager = page_doc.get('pager')
		wine_results = page_doc.get('result')
		
		# Find wine records on page
		wine_records_on_page = int(page_pager.get('current_page_record_count'))
		
		# For each record on page 
		for j in range(0, (wine_records_on_page)):
			
			name = wine_results[j].get('name')
			price = int(wine_results[j].get('price_in_cents'))
						
			# Write data to next row
			sheet.cell(row = row_iter, column = 1).value = name
			sheet.cell(row = row_iter, column = 2).value = price
				
			# Increment row_iter
			row_iter += 1
		
	wb.save('wineData.xlsx')
	
	
def excel_writer():

	global row_iter
	
	for j in range(0, (wine_records_on_page)):
			
			name = wine_results[j].get('name')
			price = int(wine_results[j].get('price_in_cents'))
						
			# Write data to next row
			sheet.cell(row = row_iter, column = 1).value = name
			sheet.cell(row = row_iter, column = 2).value = price
				
			# Increment row_iter
			row_iter += 1

def wine_score_api():
	
	headers = {'Authorization': 'Token %s' % TOKEN}
	payload = {'wine': 'bordeaux'}
	
	r = requests.get(GWS_URL, payload, headers=headers)
	print(r.url)
	print(r.status_code)	
	
lcbo_api()
#wine_score_api()
